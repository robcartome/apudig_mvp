"""
inventory/views/reports.py — Vistas de reportes de inventario.

Rutas:
  stock_report_detail        GET  /inventory/reportes/stock-almacen/
  stock_comparative_report   GET  /inventory/reportes/stock-comparativo/
  kardex_report              GET  /inventory/reportes/kardex/
"""
import io
from datetime import date

from django.db.models import Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from apps.partners.models import Customer, Supplier

from ..models import Category, Movement, Product
from ..selectors import (
    get_movement_traceability_report,
    get_kardex_report,
    get_stock_comparative,
    get_stock_report_enhanced,
    get_warehouses_for_store,
)


def _require_auth(request):
    if not request.user.is_authenticated:
        return redirect("login")
    return None


def _get_store_id(request):
    return getattr(request, "active_store_id", None) or request.session.get("active_store_id")


def _get_company_id(request):
    return getattr(request, "active_company_id", None) or request.session.get("active_company_id")


# ── Stock por Almacén ─────────────────────────────────────────────────────────

def stock_report_detail(request):
    """Stock por almacén con mínimo, estado, precio de compra y valorización."""
    r = _require_auth(request)
    if r:
        return r

    store_id = _get_store_id(request)
    warehouses = get_warehouses_for_store(store_id, active_only=True) if store_id else []
    selected_warehouse = request.GET.get("warehouse", "")
    search_query = request.GET.get("q", "").strip()
    fmt = request.GET.get("format", "")
    should_run = request.GET.get("run") == "1" or fmt in {"excel", "print"}
    selected_warehouse_name = ""
    if selected_warehouse:
        selected = next((w for w in warehouses if str(w.pk) == str(selected_warehouse)), None)
        selected_warehouse_name = selected.name if selected else selected_warehouse

    rows = get_stock_report_enhanced(store_id, selected_warehouse, search_query) if (store_id and should_run) else []

    # Aggregate totals
    total_stock = sum(r["quantity"] for r in rows)
    total_valuation = sum(r["valuation"] for r in rows)

    if fmt == "excel":
        return _stock_report_excel(rows, selected_warehouse, warehouses)
    if fmt == "print":
        return render(request, "inventory/reports/stock_report_print.html", {
            "rows": rows,
            "warehouses": warehouses,
            "selected_warehouse": selected_warehouse,
            "selected_warehouse_name": selected_warehouse_name,
            "q": search_query,
            "total_stock": total_stock,
            "total_valuation": total_valuation,
        })

    return render(request, "inventory/reports/stock_report_detail.html", {
        "rows": rows,
        "warehouses": warehouses,
        "selected_warehouse": selected_warehouse,
        "q": search_query,
        "total_stock": total_stock,
        "total_valuation": total_valuation,
        "total": len(rows),
        "should_run": should_run,
    })


def _stock_report_excel(rows, selected_warehouse, warehouses):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock por Almacén"

    header_fill = PatternFill("solid", fgColor="1A7F64")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Almacén", "SKU", "Producto", "Categoría", "UM", "Stock", "Mínimo", "Estado", "P. Compra (S/)", "Valorización (S/)"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row["warehouse"], row["sku"], row["product"], row["category"], row["unit"],
            float(row["quantity"]), float(row["min_stock"]), row["status"],
            float(row["price_purchase"]), float(row["valuation"]),
        ])

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="stock_por_almacen.xlsx"'
    return response


# ── Stock Comparativo ─────────────────────────────────────────────────────────

def stock_comparative_report(request):
    """Stock pivotado: productos como filas, almacenes como columnas."""
    r = _require_auth(request)
    if r:
        return r

    store_id = _get_store_id(request)
    all_warehouses = get_warehouses_for_store(store_id, active_only=True) if store_id else []
    selected_ids = request.GET.getlist("warehouse")
    search_query = request.GET.get("q", "").strip()
    fmt = request.GET.get("format", "")
    should_run = request.GET.get("run") == "1" or fmt in {"excel", "print"}

    warehouses, rows, summary = ([], [], {"by_warehouse": {}, "grand_total": 0, "grand_valuation": 0})
    if store_id and should_run:
        warehouses, rows, summary = get_stock_comparative(store_id, selected_ids or None, search_query)

    if fmt == "excel":
        return _comparative_excel(warehouses, rows, summary)
    if fmt == "print":
        return render(request, "inventory/reports/stock_comparative_print.html", {
            "warehouses": warehouses,
            "rows": rows,
            "summary": summary,
            "all_warehouses": all_warehouses,
            "selected_ids": selected_ids,
            "q": search_query,
        })

    return render(request, "inventory/reports/stock_comparative.html", {
        "warehouses": warehouses,
        "rows": rows,
        "summary": summary,
        "all_warehouses": all_warehouses,
        "selected_ids": selected_ids,
        "q": search_query,
        "total": len(rows),
        "should_run": should_run,
    })


def _comparative_excel(warehouses, rows, summary):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Comparativo"

    header_fill = PatternFill("solid", fgColor="1A7F64")
    header_font = Font(bold=True, color="FFFFFF")

    wh_names = [wh.name for wh in warehouses]
    headers = ["SKU", "Producto", "Categoría", "UM"] + wh_names + ["Stock Total", "P. Compra (S/)", "P. Venta (S/)", "Valorización (S/)"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    wh_ids = [str(wh.id) for wh in warehouses]
    for row in rows:
        data = [row["sku"], row["product"], row["category"], row["unit"]]
        for wid in wh_ids:
            data.append(float(row["stocks"].get(wid, 0)))
        data += [
            float(row["total_stock"]),
            float(row["price_purchase"]),
            float(row["price_sale"]),
            float(row["total_valuation"]),
        ]
        ws.append(data)

    # Totals row
    total_row = ["", "TOTAL", "", ""]
    for wid in wh_ids:
        total_row.append(float(summary["by_warehouse"].get(wid, 0)))
    total_row += [float(summary["grand_total"]), "", "", float(summary["grand_valuation"])]
    ws.append(total_row)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col_idx).font = Font(bold=True)

    for idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(idx)
        max_len = 0
        for cell in ws[column_letter]:
            if getattr(cell, "value", None) is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="stock_comparativo.xlsx"'
    return response


# ── Kardex de Inventario ──────────────────────────────────────────────────────

def kardex_report(request):
    """Kardex de inventario con saldo anterior, ingresos, salidas y saldo."""
    r = _require_auth(request)
    if r:
        return r

    store_id = _get_store_id(request)
    company_id = _get_company_id(request)
    warehouses = get_warehouses_for_store(store_id, active_only=True) if store_id else []
    categories = Category.objects.filter(company_id=company_id, active=True).order_by("name") if company_id else []

    # Params
    warehouse_id = request.GET.get("warehouse", "")
    today = timezone.localdate()
    year = int(request.GET.get("year", today.year))
    date_from_str = request.GET.get("date_from", f"{year}-01-01")
    date_to_str = request.GET.get("date_to", today.isoformat())
    category_id = request.GET.get("category", "")
    product_search = request.GET.get("product_q", "").strip()
    show_previous = request.GET.get("show_previous", "1") != "0"
    fmt = request.GET.get("format", "")
    should_run = request.GET.get("run") == "1" or fmt in {"excel", "print"}

    try:
        date_from = date.fromisoformat(date_from_str)
        date_to = date.fromisoformat(date_to_str)
    except ValueError:
        date_from = date(year, 1, 1)
        date_to = today

    years_range = range(today.year - 5, today.year + 1)

    kardex_groups = []
    if store_id and warehouse_id and should_run:
        # Resolve product_id from search
        product_id = ""
        if product_search:
            from ..models import Product as ProductModel
            prod = (
                ProductModel.objects.filter(company_id=company_id)
                .filter(
                    Q(name__icontains=product_search)
                    | Q(sku__icontains=product_search)
                    | Q(barcode__icontains=product_search)
                )
                .first()
            )
            product_id = str(prod.pk) if prod else "__no_match__"

        kardex_groups = get_kardex_report(
            store_id=store_id,
            warehouse_id=warehouse_id,
            date_from=date_from,
            date_to=date_to,
            category_id=category_id,
            product_id=product_id,
            show_previous_balance=show_previous,
        )

    context = {
        "warehouses": warehouses,
        "categories": categories,
        "warehouse_id": warehouse_id,
        "year": year,
        "date_from": date_from_str,
        "date_to": date_to_str,
        "category_id": category_id,
        "product_q": product_search,
        "show_previous": show_previous,
        "years_range": years_range,
        "kardex_groups": kardex_groups,
        "total_products": len(kardex_groups),
        "should_run": should_run,
    }

    if fmt == "excel":
        return _kardex_excel(kardex_groups, warehouse_id, warehouses, date_from_str, date_to_str)
    if fmt == "print":
        return render(request, "inventory/reports/kardex_print.html", context)

    return render(request, "inventory/reports/kardex_report.html", context)


# ── Movimientos de Almacén ───────────────────────────────────────────────────

def movement_traceability_report(request):
    """Reporte de movimientos trazables por producto con links al detalle."""
    r = _require_auth(request)
    if r:
        return r

    store_id = _get_store_id(request)
    company_id = _get_company_id(request)
    warehouses = get_warehouses_for_store(store_id, active_only=True) if store_id else []

    selected_warehouse = request.GET.get("warehouse", "")
    movement_type = request.GET.get("type", "")
    search_query = request.GET.get("q", "").strip()
    selected_product_ids = request.GET.getlist("product")
    fmt = request.GET.get("format", "")
    should_run = request.GET.get("run") == "1" or fmt in {"excel", "print"}

    today = timezone.localdate()
    default_date_from = today.replace(month=1, day=1)
    date_from_str = request.GET.get("date_from", default_date_from.isoformat())
    date_to_str = request.GET.get("date_to", today.isoformat())

    try:
        date_from = date.fromisoformat(date_from_str)
    except ValueError:
        date_from = default_date_from
        date_from_str = date_from.isoformat()

    try:
        date_to = date.fromisoformat(date_to_str)
    except ValueError:
        date_to = today
        date_to_str = date_to.isoformat()

    selected_products = []
    if selected_product_ids and company_id:
        selected_products = list(
            Product.objects.filter(company_id=company_id, id__in=selected_product_ids)
            .select_related("unit")
            .order_by("sku")
        )

    report = {
        "products": [],
        "summary_by_type": [],
        "total_rows": 0,
        "total_quantity": 0,
        "total_amount": 0,
        "rows": [],
    }

    selected_warehouse_name = selected_warehouse
    if selected_warehouse:
        selected_wh = next((w for w in warehouses if str(w.pk) == str(selected_warehouse)), None)
        if selected_wh:
            selected_warehouse_name = selected_wh.name

    movement_type_label = next((label for code, label in Movement.MOVEMENT_TYPES if code == movement_type), movement_type)

    if store_id and should_run:
        report = get_movement_traceability_report(
            store_id=store_id,
            warehouse_id=selected_warehouse,
            movement_type=movement_type,
            date_from=date_from,
            date_to=date_to,
            product_ids=selected_product_ids or None,
            search=search_query,
            limit=5000 if fmt == "" else None,
        )

    context = {
        "warehouses": warehouses,
        "movement_types": Movement.MOVEMENT_TYPES,
        "selected_warehouse": selected_warehouse,
        "movement_type": movement_type,
        "movement_type_label": movement_type_label,
        "q": search_query,
        "date_from": date_from_str,
        "date_to": date_to_str,
        "selected_product_ids": selected_product_ids,
        "selected_products": selected_products,
        "selected_warehouse_name": selected_warehouse_name,
        "report": report,
        "total_products": len(report["products"]),
        "total_rows": report["total_rows"],
        "should_run": should_run,
        "product_search_url": reverse("inventory:api_product_search"),
    }

    if fmt == "excel":
        return _movement_traceability_excel(report, selected_warehouse_name, movement_type_label, date_from_str, date_to_str)
    if fmt == "print":
        return render(request, "inventory/reports/movement_traceability_print.html", context)

    return render(request, "inventory/reports/movement_traceability.html", context)


def _movement_traceability_excel(report, selected_warehouse_name, movement_type_label, date_from_str, date_to_str):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    header_fill = PatternFill("solid", fgColor="1A7F64")
    header_font = Font(bold=True, color="FFFFFF")
    product_fill = PatternFill("solid", fgColor="D9EAD3")
    product_font = Font(bold=True)

    ws.append(["Reporte de Movimientos de Almacén"])
    ws.append([f"Rango: {date_from_str} a {date_to_str}"])
    if selected_warehouse_name:
        ws.append([f"Almacén: {selected_warehouse_name}"])
    if movement_type_label:
        ws.append([f"Tipo: {movement_type_label}"])
    ws.append([])

    ws.append([
        "Fecha", "Tipo", "Almacén", "Producto", "SKU", "Socio", "Documento",
        "Referencia", "Cantidad", "Cantidad firmada", "P. Unit", "Importe",
    ])
    for col_idx in range(1, 13):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for product in report["products"]:
        ws.append([f"{product['sku']} - {product['product']}"])
        for col_idx in range(1, 13):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = product_fill
            cell.font = product_font
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=12)

        for entry in product["entries"]:
            ws.append([
                entry["date"].strftime("%d/%m/%Y %H:%M"),
                entry["type_label"],
                entry["warehouse"],
                entry["product"],
                entry["sku"],
                entry["partner"],
                entry["document"],
                entry["reference_doc"],
                float(entry["quantity"]),
                float(entry["signed_quantity"]),
                float(entry["unit_price"]),
                float(entry["amount"]),
            ])
        ws.append([])

    for idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(idx)
        max_len = 0
        for cell in ws[column_letter]:
            if getattr(cell, "value", None) is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 4, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="movimientos_almacen.xlsx"'
    return response


def movement_operation_readonly(request, pk):
    """Vista solo lectura de una operación/movimiento."""
    r = _require_auth(request)
    if r:
        return r

    movement = get_object_or_404(
        Movement.objects
        .select_related(
            "warehouse",
            "warehouse_origin",
            "warehouse_dest",
            "supplier",
            "customer",
            "carrier",
            "document_type",
            "created_by",
            "confirmed_by",
            "closed_by",
        )
        .prefetch_related("details__product__unit", "details__location"),
        pk=pk,
    )
    return render(request, "inventory/reports/readonly/movement_operation.html", {"movement": movement})


def product_readonly(request, pk):
    """Vista solo lectura de producto en formato tipo formulario."""
    r = _require_auth(request)
    if r:
        return r

    company_id = _get_company_id(request)
    qs = Product.objects.select_related("unit", "category", "brand")
    if company_id:
        qs = qs.filter(company_id=company_id)
    product = get_object_or_404(qs, pk=pk)
    return render(request, "inventory/reports/readonly/product.html", {"product": product})


def partner_readonly(request, kind, pk):
    """Vista solo lectura para cliente o proveedor."""
    r = _require_auth(request)
    if r:
        return r

    company_id = _get_company_id(request)
    if kind == "customer":
        qs = Customer.objects.all()
        if company_id:
            qs = qs.filter(company_id=company_id)
        partner = get_object_or_404(qs, pk=pk)
        return render(request, "inventory/reports/readonly/customer.html", {"partner": partner})

    if kind == "supplier":
        qs = Supplier.objects.all()
        if company_id:
            qs = qs.filter(company_id=company_id)
        partner = get_object_or_404(qs, pk=pk)
        return render(request, "inventory/reports/readonly/supplier.html", {"partner": partner})

    raise Http404("Tipo de socio no soportado")


def _kardex_excel(kardex_groups, warehouse_id, warehouses, date_from, date_to):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    header_fill = PatternFill("solid", fgColor="1A7F64")
    header_font = Font(bold=True, color="FFFFFF")
    product_fill = PatternFill("solid", fgColor="D9EAD3")
    product_font = Font(bold=True)

    headers = [
        "Fecha", "Tipo", "Documento", "Referencia", "Socio de Negocio",
        "Ingreso", "Salida", "Saldo",
        "Valor Unit.", "Costo Prom.", "Val. Ingreso", "Val. Salida", "Val. Saldo",
    ]

    for group in kardex_groups:
        # Product header row
        ws.append([f"[{group['sku']}] {group['product_name']} — {group['category']} ({group['unit']})"])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = product_fill
            cell.font = product_font
        ws.merge_cells(
            start_row=ws.max_row, start_column=1,
            end_row=ws.max_row, end_column=len(headers),
        )

        # Column headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for entry in group["entries"]:
            ws.append([
                entry["date"].strftime("%d/%m/%Y %H:%M") if entry["date"] else "",
                entry["type"],
                entry["document"],
                entry["reference"],
                entry["partner"],
                float(entry["ingreso"]) if entry["ingreso"] is not None else "",
                float(entry["salida"]) if entry["salida"] is not None else "",
                float(entry["saldo"]),
                float(entry["unit_value"]) if entry["unit_value"] is not None else "",
                float(entry["avg_cost"]) if entry["avg_cost"] is not None else "",
                float(entry["value_ingreso"]) if entry["value_ingreso"] is not None else "",
                float(entry["value_salida"]) if entry["value_salida"] is not None else "",
                float(entry["value_saldo"]),
            ])

        ws.append([])  # blank row between products

    for idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(idx)
        max_len = 0
        for cell in ws[column_letter]:
            if getattr(cell, "value", None) is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    suffix = f"_{date_from}_{date_to}"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="kardex{suffix}.xlsx"'
    return response
