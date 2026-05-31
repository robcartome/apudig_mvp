"""
inventory/importers.py - Bulk import helpers for inventory masters.
"""
from __future__ import annotations

import csv
import io
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from django.utils.text import slugify

from .models import Brand, Category, Product, Unit


@dataclass
class ImportResult:
    entity: str
    total_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []

    @property
    def ok(self) -> bool:
        return not self.errors


ENTITY_LABELS = {
    "categories": "categorias",
    "brands": "marcas",
    "units": "unidades",
    "products": "productos",
}

IMPORT_TEMPLATE_HEADERS: dict[str, list[str]] = {
    "categories": ["codigo", "nombre", "activo"],
    "brands": ["nombre", "activo"],
    "units": ["codigo", "nombre"],
    "products": [
        "sku",
        "nombre",
        "unidad_codigo",
        "unidad_nombre",
        "categoria_codigo",
        "categoria",
        "marca",
        "precio_compra",
        "precio_venta",
        "codigo_barras",
        "descripcion",
        "modelo",
        "activo",
    ],
}

IMPORT_TEMPLATE_SAMPLE: dict[str, dict[str, Any]] = {
    "categories": {"codigo": "FER", "nombre": "Ferreteria", "activo": "si"},
    "brands": {"nombre": "Truper", "activo": "si"},
    "units": {"codigo": "NIU", "nombre": "Unidad"},
    "products": {
        "sku": "RA05002",
        "nombre": "Manguera duplex 5/8 x 100",
        "unidad_codigo": "NIU",
        "unidad_nombre": "Unidad",
        "categoria_codigo": "FER",
        "categoria": "Ferreteria",
        "marca": "Generica",
        "precio_compra": "25.50",
        "precio_venta": "35.00",
        "codigo_barras": "",
        "descripcion": "",
        "modelo": "",
        "activo": "si",
    },
}


def get_import_template_headers(entity: str) -> list[str]:
    if entity not in IMPORT_TEMPLATE_HEADERS:
        raise ValueError("Tipo de importacion no soportado.")
    return IMPORT_TEMPLATE_HEADERS[entity]


def build_import_template_workbook(entity: str) -> bytes:
    headers = get_import_template_headers(entity)
    sample = IMPORT_TEMPLATE_SAMPLE.get(entity, {})

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValueError("Falta dependencia openpyxl. Instala openpyxl para generar plantilla.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "plantilla"

    ws.append(headers)
    ws.append([sample.get(h, "") for h in headers])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _norm(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.lower()


def _as_bool(value: Any, default: bool = True) -> bool:
    text = _norm(value)
    if text == "":
        return default
    return text in {"1", "true", "si", "yes", "y", "activo", "activa", "x"}


def _as_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value is None or str(value).strip() == "":
        return default
    text = str(value).strip().replace(" ", "")
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return default


def _read_tabular(file_obj, filename: str) -> list[dict[str, Any]]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        decoded = io.TextIOWrapper(file_obj, encoding="utf-8-sig")
        reader = csv.DictReader(decoded)
        return [dict(row) for row in reader]

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Falta dependencia openpyxl. Instala openpyxl para importar Excel.") from exc

    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[idx] if idx < len(row) else None
        if any(str(v).strip() for v in item.values() if v is not None):
            data.append(item)
    return data


def _pick(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    normalized = {_norm(k): v for k, v in row.items()}
    for alias in aliases:
        key = _norm(alias)
        if key in normalized:
            return normalized[key]
    return None


def _build_category_code(name: str) -> str:
    base = (slugify(name or "", allow_unicode=False) or "cat").upper().replace("-", "_")
    return base[:30]


def _ensure_category(company, code: str | None, name: str | None) -> Category | None:
    if not code and not name:
        return None

    if code:
        obj, _ = Category.objects.get_or_create(
            company=company,
            code=str(code).strip(),
            defaults={"name": (str(name or code).strip()[:255] or str(code).strip()), "active": True},
        )
        return obj

    raw_name = str(name).strip()
    if not raw_name:
        return None

    obj = Category.objects.filter(company=company, name__iexact=raw_name).first()
    if obj:
        return obj

    base = _build_category_code(raw_name)
    candidate = base
    suffix = 1
    while Category.objects.filter(company=company, code=candidate).exists():
        suffix += 1
        candidate = f"{base[:26]}_{suffix}"

    return Category.objects.create(company=company, code=candidate, name=raw_name, active=True)


def _ensure_brand(company, name: str | None) -> Brand | None:
    raw_name = str(name or "").strip()
    if not raw_name:
        return None
    obj, _ = Brand.objects.get_or_create(company=company, name=raw_name, defaults={"active": True})
    return obj


def _ensure_unit(code: str | None, name: str | None) -> Unit | None:
    raw_code = str(code or "").strip().upper()
    raw_name = str(name or "").strip()
    if not raw_code and not raw_name:
        return None
    if not raw_code:
        raw_code = (slugify(raw_name or "und", allow_unicode=False) or "UND").upper().replace("-", "")[:10]
    if not raw_name:
        raw_name = raw_code
    obj, _ = Unit.objects.get_or_create(code=raw_code, defaults={"name": raw_name})
    return obj


def _import_categories(rows: list[dict[str, Any]], company, result: ImportResult) -> None:
    for index, row in enumerate(rows, start=2):
        code = str(_pick(row, ("code", "codigo", "categoria_codigo")) or "").strip()
        name = str(_pick(row, ("name", "nombre", "categoria", "categoria_nombre")) or "").strip()
        active = _as_bool(_pick(row, ("active", "activo", "estado")), default=True)

        if not code:
            result.errors.append(f"Fila {index}: codigo es obligatorio.")
            continue
        if not name:
            result.errors.append(f"Fila {index}: nombre es obligatorio.")
            continue

        obj, created = Category.objects.get_or_create(
            company=company,
            code=code,
            defaults={"name": name, "active": active},
        )
        if created:
            result.created += 1
        else:
            obj.name = name
            obj.active = active
            obj.save(update_fields=["name", "active", "updated_at"])
            result.updated += 1


def _import_brands(rows: list[dict[str, Any]], company, result: ImportResult) -> None:
    for index, row in enumerate(rows, start=2):
        name = str(_pick(row, ("name", "nombre", "marca", "brand")) or "").strip()
        active = _as_bool(_pick(row, ("active", "activo", "estado")), default=True)

        if not name:
            result.errors.append(f"Fila {index}: nombre es obligatorio.")
            continue

        obj, created = Brand.objects.get_or_create(
            company=company,
            name=name,
            defaults={"active": active},
        )
        if created:
            result.created += 1
        else:
            obj.active = active
            obj.save(update_fields=["active", "updated_at"])
            result.updated += 1


def _import_units(rows: list[dict[str, Any]], result: ImportResult) -> None:
    for index, row in enumerate(rows, start=2):
        code = str(_pick(row, ("code", "codigo", "unit_code", "unidad_codigo")) or "").strip().upper()
        name = str(_pick(row, ("name", "nombre", "unit_name", "unidad_nombre", "unidad")) or "").strip()

        if not code:
            result.errors.append(f"Fila {index}: codigo es obligatorio.")
            continue
        if not name:
            result.errors.append(f"Fila {index}: nombre es obligatorio.")
            continue

        obj, created = Unit.objects.get_or_create(code=code, defaults={"name": name})
        if created:
            result.created += 1
        else:
            obj.name = name
            obj.save(update_fields=["name"])
            result.updated += 1


def _import_products(rows: list[dict[str, Any]], company, result: ImportResult) -> None:
    for index, row in enumerate(rows, start=2):
        sku = str(_pick(row, ("sku", "codigo", "codigo_producto")) or "").strip()
        name = str(_pick(row, ("name", "nombre", "producto", "producto_nombre")) or "").strip()
        barcode = str(_pick(row, ("barcode", "codigo_barras", "ean")) or "").strip()
        description = str(_pick(row, ("description", "descripcion")) or "").strip()
        model = str(_pick(row, ("model", "modelo")) or "").strip()
        active = _as_bool(_pick(row, ("active", "activo", "estado")), default=True)

        unit_code = str(_pick(row, ("unit_code", "unidad_codigo", "unit", "unidad")) or "").strip()
        unit_name = str(_pick(row, ("unit_name", "unidad_nombre")) or "").strip()
        category_code = str(_pick(row, ("category_code", "categoria_codigo")) or "").strip()
        category_name = str(_pick(row, ("category", "categoria", "categoria_nombre")) or "").strip()
        brand_name = str(_pick(row, ("brand", "marca", "brand_name")) or "").strip()

        if not sku:
            result.errors.append(f"Fila {index}: sku es obligatorio.")
            continue
        if not name:
            result.errors.append(f"Fila {index}: nombre es obligatorio.")
            continue

        unit = _ensure_unit(unit_code, unit_name)
        if not unit:
            result.errors.append(f"Fila {index}: unidad no valida.")
            continue

        category = _ensure_category(company, category_code, category_name)
        brand = _ensure_brand(company, brand_name)

        defaults = {
            "name": name,
            "barcode": barcode,
            "description": description,
            "model": model,
            "price_purchase": _as_decimal(_pick(row, ("price_purchase", "precio_compra", "costo"))),
            "price_sale": _as_decimal(_pick(row, ("price_sale", "precio_venta", "pvp"))),
            "category": category,
            "brand": brand,
            "unit": unit,
            "active": active,
        }

        obj, created = Product.objects.get_or_create(
            company=company,
            sku=sku,
            defaults=defaults,
        )
        if created:
            result.created += 1
        else:
            for field, val in defaults.items():
                setattr(obj, field, val)
            obj.save(update_fields=[
                "name",
                "barcode",
                "description",
                "model",
                "price_purchase",
                "price_sale",
                "category",
                "brand",
                "unit",
                "active",
                "updated_at",
            ])
            result.updated += 1


def import_inventory_excel(*, entity: str, file_obj, filename: str, company=None, dry_run: bool = False) -> ImportResult:
    if entity not in ENTITY_LABELS:
        raise ValueError("Tipo de importacion no soportado.")

    rows = _read_tabular(file_obj, filename)
    result = ImportResult(entity=entity, total_rows=len(rows))

    if entity in {"categories", "brands", "products"} and company is None:
        raise ValueError("Se requiere empresa activa para esta importacion.")

    with transaction.atomic():
        if entity == "categories":
            _import_categories(rows, company, result)
        elif entity == "brands":
            _import_brands(rows, company, result)
        elif entity == "units":
            _import_units(rows, result)
        elif entity == "products":
            _import_products(rows, company, result)

        result.skipped = max(result.total_rows - (result.created + result.updated + len(result.errors)), 0)
        if dry_run or result.errors:
            transaction.set_rollback(True)

    return result


# ── Price-list bulk helpers ───────────────────────────────────────────────────

_FIXED_PRICE_COLS = {
    "sku", "nombre", "name",
    "precio_compra", "price_purchase", "costo",
    "precio_venta_general", "precio_venta", "price_sale", "pvp",
    "codigo_barras", "barcode",
}


def build_pricelist_template_workbook(company, pricelist=None) -> bytes:
    """Build an Excel file with all active products as rows.

    Fixed columns: sku | nombre | precio_compra | precio_venta_general
    Dynamic columns: one per active PriceList (filtered to *pricelist* if given).
    A hidden 'metadata' sheet lists list name → UUID for round-trip matching.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ValueError("Falta dependencia openpyxl. Instala openpyxl para generar plantilla.") from exc

    from .models import PriceList, ProductPrice

    if pricelist is not None:
        price_lists = [pricelist]
    else:
        price_lists = list(
            PriceList.objects.filter(company=company, active=True).order_by("name")
        )

    fixed_headers = ["sku", "nombre", "precio_compra", "precio_venta_general"]
    pl_headers = [pl.name for pl in price_lists]
    all_headers = fixed_headers + pl_headers

    products = list(
        Product.objects.filter(company=company, active=True).order_by("sku")
    )

    # Build lookup: (str(product_id), str(pricelist_id)) → amount
    prices_qs = ProductPrice.objects.filter(
        product__company=company,
        price_list__in=price_lists,
    ).values("product_id", "price_list_id", "amount")
    price_map = {
        (str(p["product_id"]), str(p["price_list_id"])): p["amount"]
        for p in prices_qs
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "precios"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2B5AB7")
    lock_fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
    center = Alignment(horizontal="center")

    ws.append(all_headers)
    for col_idx in range(1, len(all_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for product in products:
        row_data = [
            product.sku,
            product.name,
            float(product.price_purchase),
            float(product.price_sale),
        ]
        for pl in price_lists:
            key = (str(product.pk), str(pl.pk))
            raw = price_map.get(key)
            row_data.append(float(raw) if raw is not None else "")
        ws.append(row_data)
        row_num = ws.max_row
        # Gray out read-only info columns
        for col_idx in (1, 2):
            ws.cell(row=row_num, column=col_idx).fill = lock_fill

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    for i in range(len(pl_headers)):
        col_letter = ws.cell(row=1, column=5 + i).column_letter
        ws.column_dimensions[col_letter].width = 22

    # Metadata sheet for UUID → name round-trip
    ws_meta = wb.create_sheet("metadata")
    ws_meta.append(["lista_nombre", "lista_id"])
    for pl in price_lists:
        ws_meta.append([pl.name, str(pl.pk)])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def import_pricelist_excel(
    *,
    file_obj,
    filename: str,
    company,
    pricelist=None,
) -> ImportResult:
    """Update product base prices and/or price-list amounts from Excel.

    Columns expected:
      sku | nombre (ignored) | precio_compra | precio_venta_general
      + one column per price list (matched by name, case-insensitive).

    If *pricelist* is provided only that list's column is processed.
    Rows with errors are skipped; valid rows are always committed.
    """
    from .models import PriceList, ProductPrice

    rows = _read_tabular(file_obj, filename)
    result = ImportResult(entity="price_lists", total_rows=len(rows))

    if not rows:
        return result

    # Build company price-list map: normalised_name → PriceList
    company_pl_map: dict[str, "PriceList"] = {
        _norm(pl.name): pl
        for pl in PriceList.objects.filter(company=company)
    }

    # Determine which extra columns map to a PriceList
    sample_keys = list(rows[0].keys())
    pl_column_map: dict[str, "PriceList"] = {}  # original column name → PriceList
    for col in sample_keys:
        if _norm(col) in _FIXED_PRICE_COLS:
            continue
        pl = company_pl_map.get(_norm(col))
        if pl is None:
            continue
        if pricelist is not None and str(pl.pk) != str(pricelist.pk):
            continue
        pl_column_map[col] = pl

    # If caller passed a specific pricelist but no matching column found, try fuzzy
    if pricelist is not None and not pl_column_map:
        for col in sample_keys:
            if _norm(col) == _norm(pricelist.name):
                pl_column_map[col] = pricelist
                break

    for index, row in enumerate(rows, start=2):
        sku = str(_pick(row, ("sku", "codigo", "codigo_articulo")) or "").strip()
        if not sku:
            result.errors.append(f"Fila {index}: SKU es obligatorio.")
            continue

        try:
            product = Product.objects.get(company=company, sku=sku)
        except Product.DoesNotExist:
            result.errors.append(f"Fila {index}: SKU '{sku}' no encontrado.")
            continue

        # Update base prices on Product
        pp_raw = _pick(row, ("precio_compra", "price_purchase", "costo"))
        ps_raw = _pick(row, ("precio_venta_general", "precio_venta", "price_sale", "pvp"))
        update_fields: list[str] = []
        if pp_raw is not None and str(pp_raw).strip() != "":
            product.price_purchase = _as_decimal(pp_raw)
            update_fields.append("price_purchase")
        if ps_raw is not None and str(ps_raw).strip() != "":
            product.price_sale = _as_decimal(ps_raw)
            update_fields.append("price_sale")
        if update_fields:
            update_fields.append("updated_at")
            product.save(update_fields=update_fields)

        # Update ProductPrice rows
        for col_name, pl in pl_column_map.items():
            amount_raw = row.get(col_name)
            if amount_raw is None or str(amount_raw).strip() == "":
                continue
            amount = _as_decimal(amount_raw)
            pp_obj, created = ProductPrice.objects.get_or_create(
                product=product,
                price_list=pl,
                defaults={"amount": amount, "currency": "PEN", "active": True},
            )
            if not created and pp_obj.amount != amount:
                pp_obj.amount = amount
                pp_obj.save(update_fields=["amount"])

        result.updated += 1

    result.skipped = max(
        result.total_rows - (result.created + result.updated + len(result.errors)), 0
    )
    return result


# ── Price report workbook ─────────────────────────────────────────────────────

def build_price_report_workbook(
    price_lists: list,
    rows: list[dict],
    show_compra: bool = True,
    show_venta: bool = True,
) -> bytes:
    """Build an Excel price report from pre-computed selector data.

    price_lists  — list of PriceList objects to include.
    rows         — list of dicts with keys: sku, name, price_purchase, price_sale,
                   prices {str(pl_pk): Decimal|None}.
    show_compra  — include Precio Compra column.
    show_venta   — include Precio Venta column.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("Falta dependencia openpyxl.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "consolidado_precios"

    headers = ["SKU", "Nombre"]
    if show_compra:
        headers.append("Precio Compra")
    if show_venta:
        headers.append("Precio Venta")
    headers += [pl.name for pl in price_lists]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2B5AB7")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in rows:
        row_data: list = [row["sku"], row["name"]]
        if show_compra:
            row_data.append(float(row["price_purchase"]) if row["price_purchase"] is not None else "")
        if show_venta:
            row_data.append(float(row["price_sale"]) if row["price_sale"] is not None else "")
        for pl in price_lists:
            val = row["prices"].get(str(pl.pk))
            row_data.append(float(val) if val is not None else "")
        ws.append(row_data)
        row_num = ws.max_row
        for col_idx in range(3, len(headers) + 1):
            ws.cell(row=row_num, column=col_idx).alignment = right

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 45
    for i in range(2, len(headers)):
        col_letter = get_column_letter(3 + i - 2 + 1)
        ws.column_dimensions[get_column_letter(3 + i - 2)].width = 18

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ── Consolidated price report Excel builder ───────────────────────────────────

def build_price_report_workbook(price_lists: list, rows: list[dict]) -> bytes:
    """Generate a styled Excel workbook from pre-built consolidate data.

    price_lists — list of PriceList objects (ordered).
    rows        — list of dicts as returned by get_price_consolidate().
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, numbers
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("Falta dependencia openpyxl.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "reporte_precios"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2B5AB7")
    sku_fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    num_fmt = "#,##0.00"

    headers = (
        ["SKU", "Nombre", "Precio Compra", "Precio Venta General"]
        + [pl.name for pl in price_lists]
    )
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in rows:
        row_data = [
            row["sku"],
            row["name"],
            float(row["price_purchase"]),
            float(row["price_sale"]),
        ] + [
            float(row["prices"].get(str(pl.pk))) if row["prices"].get(str(pl.pk)) is not None else ""
            for pl in price_lists
        ]
        ws.append(row_data)
        data_row = ws.max_row
        ws.cell(data_row, 1).fill = sku_fill
        for col_idx in range(3, len(headers) + 1):
            c = ws.cell(data_row, col_idx)
            c.number_format = num_fmt
            c.alignment = right

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    for i in range(len(price_lists)):
        col_letter = get_column_letter(5 + i)
        ws.column_dimensions[col_letter].width = 22

    ws.freeze_panes = "C2"  # freeze SKU+Nombre columns and header row

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
