"""
sales/views/pdf.py — Vista PDF de cotizaciones (HTML print-ready).

Genera una página HTML sin el layout base, optimizada para impresión A4.
El navegador la puede imprimir como PDF con Ctrl+P / window.print().
"""
from io import BytesIO

from django.http import Http404, HttpResponse
from django.utils.text import slugify
from django.http import Http404
from django.shortcuts import render

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.sales.models import SalesQuotation
from apps.sales.selectors import get_quotation_detail, search_quotations


def _require_auth(request):
    if not request.user.is_authenticated:
        from django.shortcuts import redirect
        return redirect("login")
    return None


def quotation_pdf(request, pk):
    redirect_resp = _require_auth(request)
    if redirect_resp:
        return redirect_resp

    try:
        quotation = get_quotation_detail(pk)
    except SalesQuotation.DoesNotExist:
        raise Http404

    # Obtener la empresa/compañía para el encabezado del PDF
    company = None
    if quotation.store:
        company = quotation.store.company

    return render(request, "sales/pdf/quotation_pdf.html", {
        "quotation": quotation,
        "company": company,
    })


def quotation_xlsx(request):
    redirect_resp = _require_auth(request)
    if redirect_resp:
        return redirect_resp

    company_id = getattr(request, "active_company_id", None) or request.session.get("active_company_id")
    store_id = getattr(request, "active_store_id", None) or request.session.get("active_store_id")
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()

    qs = search_quotations(store_id, query=query or None, status=status or None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"

    headers = [
        "Fecha emisión",
        "Serie",
        "Número",
        "Cliente",
        "Forma de pago",
        "Medio de pago",
        "Válido hasta",
        "Moneda",
        "Total",
        "Estado",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F3C88")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for quotation in qs:
        ws.append([
            quotation.issue_date,
            quotation.series_code,
            quotation.number,
            quotation.customer_legal_name,
            str(quotation.payment_method) if quotation.payment_method else "",
            str(quotation.means_of_payment) if quotation.means_of_payment else "",
            quotation.valid_until,
            quotation.currency,
            float(quotation.total),
            quotation.get_status_display(),
        ])

    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 35)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"cotizaciones_{slugify(query) if query else 'lista'}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
